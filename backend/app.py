from __future__ import annotations

import io
from typing import Any

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse
from matplotlib.figure import Figure
from openpyxl import load_workbook
from pydantic import BaseModel

app = FastAPI(title='Ashby Backend API')


class RenderPlotRequest(BaseModel):
    config: dict[str, Any]


def _extract_columns_from_xlsx(file_bytes: bytes, sheet_index: int) -> list[str]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = workbook.worksheets
    if not sheets:
        return []
    index = min(max(sheet_index, 0), len(sheets) - 1)
    first_row = next(sheets[index].iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(cell).strip() for cell in first_row if isinstance(cell, str) and cell.strip()]


@app.post('/api/render-plot')
def render_plot(payload: RenderPlotRequest) -> JSONResponse:
    figure = Figure(figsize=(12, 6.75), dpi=100)
    axis = figure.subplots()
    axis.plot([], [])
    axis.set_title('Backend matplotlib dummy plot')
    axis.text(
        0.01,
        0.95,
        f"Received config keys: {', '.join(sorted(payload.config.keys()))}",
        transform=axis.transAxes,
        fontsize=10,
        va='top',
    )
    axis.grid(True, alpha=0.25)

    svg_buffer = io.StringIO()
    figure.savefig(svg_buffer, format='svg', bbox_inches='tight')
    svg_buffer.seek(0)
    return JSONResponse({'svg': svg_buffer.getvalue()})


@app.post('/api/import-database')
async def import_database(
    request: Request,
    import_sheet: int = Form(0),
    file: UploadFile | None = File(None),
    teable_url: str | None = Form(None),
    API_Key: str | None = Form(None),
) -> JSONResponse:
    if request.headers.get('content-type', '').startswith('application/json'):
        payload = await request.json()
        teable_url_json = payload.get('teable_url')
        api_key_json = payload.get('API_Key')
        if not teable_url_json or not api_key_json:
            return JSONResponse({'success': False, 'message': 'Missing teable_url or API_Key.'}, status_code=400)
        return JSONResponse({'success': True, 'columns': []})

    if file is None:
        if not teable_url or not API_Key:
            return JSONResponse({'success': False, 'message': 'Missing teable_url or API_Key.'}, status_code=400)
        return JSONResponse({'success': True, 'columns': []})

    file_bytes = await file.read()
    columns = _extract_columns_from_xlsx(file_bytes, import_sheet)
    return JSONResponse({'success': True, 'columns': columns})
